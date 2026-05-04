"""
Executive Dashboard — Streamlit Web App
Upload a multi-entity P&L workbook → view an interactive dark-theme dashboard
→ area-by-area breakdowns → scenario analysis → download the completed Excel file.
"""

import streamlit as st
import pandas as pd
import numpy as np
import plotly.graph_objects as go
import plotly.express as px
from plotly.subplots import make_subplots
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io
import re

st.set_page_config(
    page_title="Executive Dashboard",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ── Design system (matches skill exactly) ────────────────────────────────────
COLORS = {
    "bg_primary":   "#1E1E2E",
    "bg_header":    "#353548",
    "bg_odd":       "#2A2A3C",
    "bg_even":      "#323245",
    "bg_kpi":       "#2A2A3E",
    "text_primary": "#FFFFFF",
    "text_secondary":"#E0E0E0",
    "text_muted":   "#B0B0B0",
    "text_dim":     "#78909C",
    "accent":       "#4FC3F7",
    "accent_alt":   "#42A5F5",
    "positive":     "#66BB6A",
    "negative":     "#EF5350",
    "warning":      "#FFB74D",
    "highlight":    "#FFD600",
}
CHART_COLORS = ["#78909C", "#90A4AE", "#4FC3F7", "#42A5F5", "#66BB6A", "#FF8A65"]
AREA_COLORS  = ["#4FC3F7", "#66BB6A", "#FFB74D", "#EF5350", "#AB47BC",
                 "#FF8A65", "#26C6DA", "#9CCC65"]


# ── CSS injection ─────────────────────────────────────────────────────────────
def inject_css():
    st.markdown(f"""
    <style>
    .stApp {{ background-color: {COLORS["bg_primary"]}; }}
    .main .block-container {{ padding-top: 1rem; max-width: 1400px; }}
    h1, h2, h3 {{ color: {COLORS["accent"]} !important; }}
    p, span, label, .stMarkdown {{ color: {COLORS["text_secondary"]} !important; }}
    .kpi-card {{
        background: {COLORS["bg_kpi"]}; border-radius: 8px;
        padding: 16px 20px; text-align: center; min-height: 110px;
    }}
    .kpi-label {{
        color: {COLORS["text_muted"]}; font-size: 12px; font-weight: 600;
        text-transform: uppercase; letter-spacing: 0.5px; margin-bottom: 6px;
    }}
    .kpi-value {{ font-size: 28px; font-weight: 700; color: {COLORS["text_primary"]}; line-height: 1.2; }}
    .kpi-sub   {{ color: {COLORS["text_dim"]}; font-size: 11px; margin-top: 4px; }}
    .section-title {{
        color: {COLORS["accent"]} !important; font-size: 16px; font-weight: 700;
        margin: 30px 0 12px 0; letter-spacing: 0.5px;
    }}
    .area-card {{
        background: {COLORS["bg_odd"]}; border-radius: 8px;
        padding: 14px 18px; min-height: 90px;
        border-left: 4px solid {COLORS["accent"]};
    }}
    .area-name   {{ color: {COLORS["accent"]}; font-size: 13px; font-weight: 700; margin-bottom: 4px; }}
    .area-metric {{ color: {COLORS["text_primary"]}; font-size: 20px; font-weight: 700; line-height: 1.3; }}
    .area-sub    {{ color: {COLORS["text_dim"]}; font-size: 10px; }}
    .stTabs [data-baseweb="tab-list"] {{ gap: 8px; }}
    .stTabs [data-baseweb="tab"] {{
        background-color: {COLORS["bg_header"]}; color: {COLORS["text_muted"]};
        border-radius: 6px 6px 0 0; padding: 8px 20px;
    }}
    .stTabs [aria-selected="true"] {{
        background-color: {COLORS["accent_alt"]} !important; color: white !important;
    }}
    </style>
    """, unsafe_allow_html=True)


# ── P&L parsing ───────────────────────────────────────────────────────────────
def find_row_by_label(df, keywords, col_idx=0):
    for i, val in enumerate(df.iloc[:, col_idx]):
        if pd.isna(val):
            continue
        val_str = str(val).strip().lower()
        if any(kw.lower() in val_str for kw in keywords):
            return i
    return None

def find_fy_columns(df):
    fy_cols = {}
    for row_idx in range(min(6, len(df))):
        for col_idx in range(len(df.columns)):
            val = df.iloc[row_idx, col_idx]
            if pd.isna(val):
                continue
            val_str = str(val).strip()
            match = re.search(r'(?:FY\s*)?(\d{4})\s*(?:Total|Budget|Bgt|Actual|Act)', val_str, re.IGNORECASE)
            if match and "total" in val_str.lower():
                fy_cols[f"FY{match.group(1)}"] = col_idx
    return fy_cols

def find_monthly_columns(fy_total_col):
    if fy_total_col is None or fy_total_col < 12:
        return []
    return list(range(fy_total_col - 12, fy_total_col))

def discover_and_extract(df):
    rows = {
        "total_revenue": find_row_by_label(df, ["total revenue", "total rev", "net revenue"]),
        "total_cos":     find_row_by_label(df, ["total cost of services", "total cos", "cost of services", "total cogs", "cost of goods"]),
        "total_gm":      find_row_by_label(df, ["total gross margin", "gross margin", "gross profit"]),
        "gm_pct":        find_row_by_label(df, ["gross margin %", "gm %", "gm%", "gross margin percent"]),
        "emp_comp":      find_row_by_label(df, ["employee compensation", "employee comp", "salaries", "payroll"]),
        "rent":          find_row_by_label(df, ["rent & facilities", "rent", "facilities", "occupancy"]),
        "other_direct":  find_row_by_label(df, ["other direct", "other operating"]),
        "total_direct":  find_row_by_label(df, ["total direct costs", "total direct", "total operating expenses"]),
        "cm":            find_row_by_label(df, ["contribution margin"]),
        "cm_pct":        find_row_by_label(df, ["contribution margin %", "cm %", "cm%"]),
        "oi":            find_row_by_label(df, ["operating income", "net income", "ebitda", "ebit"]),
    }
    fy_cols = find_fy_columns(df)
    annual = {}
    for fy, col in sorted(fy_cols.items()):
        a = {}
        for key, row_idx in rows.items():
            if row_idx is not None and col < len(df.columns):
                val = df.iloc[row_idx, col]
                a[key] = float(val) if pd.notna(val) and isinstance(val, (int, float)) else 0
            else:
                a[key] = 0
        annual[fy] = a

    # Monthly data
    monthly_rev, monthly_gm, monthly_cm = {}, {}, {}
    for fy, col in sorted(fy_cols.items()):
        mc = find_monthly_columns(col)
        if not mc:
            continue
        for metric_key, row_key, target in [
            ("monthly_revenue", "total_revenue", monthly_rev),
            ("monthly_gm_pct",  "gm_pct",        monthly_gm),
            ("monthly_cm_pct",  "cm_pct",         monthly_cm),
        ]:
            ri = rows.get(row_key)
            if ri is not None:
                vals = []
                for c in mc:
                    v = df.iloc[ri, c] if c < len(df.columns) else 0
                    vals.append(float(v) if pd.notna(v) and isinstance(v, (int, float)) else 0)
                target[fy] = vals

    # Service lines
    service_lines = []
    if rows["total_revenue"] is not None:
        for i in range(max(0, rows["total_revenue"] - 15), rows["total_revenue"]):
            val = df.iloc[i, 0]
            if pd.isna(val):
                continue
            name = str(val).strip()
            if not name or "total" in name.lower():
                continue
            has_data = any(
                pd.notna(df.iloc[i, c]) and isinstance(df.iloc[i, c], (int, float))
                for c in range(1, min(len(df.columns), 20))
            )
            if not has_data:
                continue
            sl = {"name": name}
            for fy, col in sorted(fy_cols.items()):
                v = df.iloc[i, col] if col < len(df.columns) else 0
                sl[fy] = float(v) if pd.notna(v) and isinstance(v, (int, float)) else 0
            service_lines.append(sl)

    return {
        "annual": annual,
        "monthly_revenue": monthly_rev,
        "monthly_gm_pct":  monthly_gm,
        "monthly_cm_pct":  monthly_cm,
        "service_lines": service_lines,
        "fy_labels": sorted(fy_cols.keys()),
        "month_labels": ["Jul","Aug","Sep","Oct","Nov","Dec","Jan","Feb","Mar","Apr","May","Jun"],
    }

def discover_area_sheets(sheet_names, consolidated_name):
    skip = ["consolidated","dashboard","fte","supervision","chart","template",
            "summary","discontinued","new sites","_chartdata","assumption"]
    areas = []
    for name in sheet_names:
        n = name.lower()
        if n == consolidated_name.lower():
            continue
        if any(s in n for s in skip):
            continue
        if "p&l" in n or "p_l" in n or "pl" in n:
            areas.append(name)
    return areas


# ── Chart helpers ─────────────────────────────────────────────────────────────
def plotly_base(title="", height=350):
    return dict(
        title=dict(text=title, font=dict(color=COLORS["text_secondary"], size=13)),
        paper_bgcolor=COLORS["bg_primary"],
        plot_bgcolor=COLORS["bg_primary"],
        font=dict(family="Segoe UI, sans-serif", color=COLORS["text_muted"], size=11),
        legend=dict(font=dict(color=COLORS["text_muted"], size=10),
                    orientation="h", yanchor="bottom", y=-0.28, xanchor="center", x=0.5),
        margin=dict(l=50, r=20, t=45, b=60),
        height=height,
        xaxis=dict(gridcolor="#3A3A50", zerolinecolor="#3A3A50"),
        yaxis=dict(gridcolor="#3A3A50", zerolinecolor="#3A3A50"),
    )

def fmt_dollars(v):
    """Format value (raw, not /1000) as $#,##0 display."""
    if v is None or (isinstance(v, float) and np.isnan(v)):
        return "—"
    return f"${abs(v)/1000:,.0f}"

def fmt_pct(v):
    if v is None or (isinstance(v, float) and np.isnan(v)):
        return "—"
    p = v * 100 if abs(v) < 2 else v
    return f"{p:.1f}%"

def fmt_growth(v):
    if v is None or (isinstance(v, float) and np.isnan(v)):
        return "—"
    p = v * 100 if abs(v) < 2 else v
    sign = "+" if p >= 0 else ""
    return f"{sign}{p:.1f}%"


# ── KPI card HTML ─────────────────────────────────────────────────────────────
def kpi_card(label, value, sub, border_color):
    return f"""
    <div class="kpi-card" style="border-top: 3px solid {border_color};">
        <div class="kpi-label">{label}</div>
        <div class="kpi-value">{value}</div>
        <div class="kpi-sub">{sub}</div>
    </div>"""

def area_card(name, revenue, cm_pct, growth, color):
    g_color = COLORS["positive"] if growth >= 0 else COLORS["negative"]
    sign = "+" if growth >= 0 else ""
    return f"""
    <div class="area-card" style="border-left-color: {color};">
        <div class="area-name">{name}</div>
        <div class="area-metric">{fmt_dollars(revenue)}</div>
        <div class="area-sub">
            CM: {fmt_pct(cm_pct)} &nbsp;|&nbsp;
            <span style="color:{g_color}">Growth: {sign}{growth:.1f}%</span>
        </div>
    </div>"""


# ── Scenario math ─────────────────────────────────────────────────────────────
def compute_scenario(annual_latest, rev_adj, cos_adj):
    rev    = annual_latest.get("total_revenue", 0) * (1 + rev_adj)
    cos_pct = (annual_latest.get("total_cos", 0) / max(annual_latest.get("total_revenue", 1), 1)) + cos_adj
    cos    = rev * cos_pct
    gm     = rev - cos
    cm     = gm - annual_latest.get("total_direct", 0)
    return {
        "revenue": rev / 1000,
        "gm":      gm / 1000,
        "cm":      cm / 1000,
        "cm_pct":  cm / rev if rev else 0,
    }


# ── Excel export ──────────────────────────────────────────────────────────────
def generate_excel(data, area_datasets, original_bytes):
    wb = openpyxl.load_workbook(io.BytesIO(original_bytes))
    if "Dashboard" in wb.sheetnames:
        del wb["Dashboard"]
    ws = wb.create_sheet("Dashboard", 0)

    fills = {
        "bg":     PatternFill(start_color="1E1E2E", fill_type="solid"),
        "header": PatternFill(start_color="353548", fill_type="solid"),
        "odd":    PatternFill(start_color="2A2A3C", fill_type="solid"),
        "even":   PatternFill(start_color="323245", fill_type="solid"),
        "kpi":    PatternFill(start_color="2A2A3E", fill_type="solid"),
    }
    fonts = {
        "title":    Font(name="Segoe UI", size=16, bold=True,  color="4FC3F7"),
        "section":  Font(name="Segoe UI", size=12, bold=True,  color="4FC3F7"),
        "header":   Font(name="Segoe UI", size=9,  bold=True,  color="4FC3F7"),
        "data":     Font(name="Segoe UI", size=10, color="FFFFFF"),
        "dim":      Font(name="Segoe UI", size=9,  color="78909C"),
        "kpi_val":  Font(name="Segoe UI", size=14, bold=True,  color="FFFFFF"),
        "bold":     Font(name="Segoe UI", size=10, bold=True,  color="FFFFFF"),
        "positive": Font(name="Segoe UI", size=10, bold=True,  color="66BB6A"),
        "negative": Font(name="Segoe UI", size=10, color="EF5350"),
    }
    border_accent = Border(top=Side(style="medium", color="4FC3F7"))
    center = Alignment(horizontal="center", vertical="center")
    right  = Alignment(horizontal="right",  vertical="center")

    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 28
    for col in "CDEFGH":
        ws.column_dimensions[col].width = 16

    for row in range(1, 130):
        for col in range(1, 9):
            ws.cell(row=row, column=col).fill = fills["bg"]

    fy_list   = sorted(data["annual"].keys())
    latest_fy = fy_list[-1] if fy_list else "FY2026"
    prior_fy  = fy_list[-2] if len(fy_list) >= 2 else fy_list[0]
    latest    = data["annual"].get(latest_fy, {})
    prior     = data["annual"].get(prior_fy,  {})

    # Title
    ws["B2"] = "EXECUTIVE DASHBOARD"
    ws["B2"].font = fonts["title"]
    ws["B3"] = "($000s)"
    ws["B3"].font = fonts["dim"]

    # KPI row
    kpi_items = [
        ("Total Revenue",        latest.get("total_revenue",0)/1000, f"PY: {fmt_dollars(prior.get('total_revenue',0))}",  "$#,##0"),
        ("Rev Growth YoY",       (latest.get("total_revenue",0)/max(prior.get("total_revenue",1),1))-1,
                                 f"PY GM%: {fmt_pct(prior.get('gm_pct',0))}",                                             "0.0%"),
        ("Gross Margin %",       latest.get("gm_pct",0),             f"PY: {fmt_pct(prior.get('gm_pct',0))}",             "0.0%"),
        ("Contribution Margin",  latest.get("cm",0)/1000,            f"PY: {fmt_dollars(prior.get('cm',0))}",             "$#,##0"),
        ("CM %",                 latest.get("cm_pct",0),             f"PY: {fmt_pct(prior.get('cm_pct',0))}",             "0.0%"),
    ]
    for i, (label, value, sub, fmt) in enumerate(kpi_items):
        col = i + 3
        lbl = ws.cell(row=4, column=col, value=label)
        lbl.font = fonts["dim"]; lbl.fill = fills["kpi"]; lbl.alignment = center
        val = ws.cell(row=5, column=col, value=value)
        val.font = fonts["kpi_val"]; val.fill = fills["kpi"]
        val.alignment = center; val.number_format = fmt
        sub_cell = ws.cell(row=6, column=col, value=sub)
        sub_cell.font = fonts["dim"]; sub_cell.fill = fills["kpi"]; sub_cell.alignment = center
        ws.cell(row=7, column=col).fill = fills["kpi"]

    # Annual financial summary
    ws.cell(row=10, column=2, value="ANNUAL FINANCIAL SUMMARY").font = fonts["section"]
    ws.cell(row=12, column=2, value=" ").font = fonts["header"]
    ws.cell(row=12, column=2).fill = fills["header"]
    for i, fy in enumerate(fy_list[-5:]):
        c = ws.cell(row=12, column=3+i, value=fy)
        c.font = fonts["header"]; c.fill = fills["header"]; c.alignment = center

    metrics = [
        ("Total Revenue",       "total_revenue", "$#,##0", False),
        ("Cost of Services",    "total_cos",     "$#,##0", False),
        ("Gross Margin",        "total_gm",      "$#,##0", True),
        ("GM %",                "gm_pct",        "0.0%",   False),
        ("Direct Costs",        "total_direct",  "$#,##0", False),
        ("Contribution Margin", "cm",            "$#,##0", True),
        ("CM %",                "cm_pct",        "0.0%",   False),
        ("Operating Income",    "oi",            "$#,##0", True),
    ]
    for j, (label, key, fmt, is_sub) in enumerate(metrics):
        row  = 13 + j
        fill = fills["odd"] if j % 2 == 0 else fills["even"]
        lbl  = ws.cell(row=row, column=2, value=label)
        lbl.font = fonts["bold"] if is_sub else fonts["data"]
        lbl.fill = fill
        if is_sub:
            lbl.border = border_accent
        for i, fy in enumerate(fy_list[-5:]):
            raw = data["annual"].get(fy, {}).get(key, 0)
            val_display = raw if "pct" in key else raw / 1000
            c = ws.cell(row=row, column=3+i, value=val_display)
            c.font  = fonts["bold"] if is_sub else fonts["data"]
            c.fill  = fill; c.number_format = fmt; c.alignment = right
            if is_sub:
                c.border = border_accent

    # Area performance
    if area_datasets:
        ar = 23
        ws.cell(row=ar, column=2, value="AREA PERFORMANCE").font = fonts["section"]
        hr = ar + 2
        for i, h in enumerate(["Area", f"Rev {prior_fy}", f"Rev {latest_fy}",
                                "Growth %", f"GM% {latest_fy}", f"CM% {latest_fy}", "CAGR"]):
            c = ws.cell(row=hr, column=2+i, value=h)
            c.font = fonts["header"]; c.fill = fills["header"]; c.alignment = center
        for j, ad in enumerate(area_datasets):
            row  = hr + 1 + j
            fill = fills["odd"] if j % 2 == 0 else fills["even"]
            ap   = ad["annual"].get(prior_fy,  {})
            al   = ad["annual"].get(latest_fy, {})
            rp   = ap.get("total_revenue", 0) / 1000
            rl   = al.get("total_revenue", 0) / 1000
            growth = (rl / rp - 1) if rp else 0
            first_fy = sorted(ad["annual"].keys())[0] if ad["annual"] else prior_fy
            rf   = ad["annual"].get(first_fy, {}).get("total_revenue", 0) / 1000
            n    = max(1, int(latest_fy[2:]) - int(first_fy[2:]))
            cagr = (rl / rf) ** (1 / n) - 1 if rf > 0 else 0
            row_data = [
                (ad["display_name"], None,    fonts["data"]),
                (rp,                 "$#,##0", fonts["data"]),
                (rl,                 "$#,##0", fonts["data"]),
                (growth,             "0.0%",   fonts["positive"] if growth > 0 else fonts["negative"]),
                (al.get("gm_pct",0), "0.0%",   fonts["data"]),
                (al.get("cm_pct",0), "0.0%",   fonts["data"]),
                (cagr,               "0.0%",   fonts["positive"] if cagr > 0.15 else fonts["data"]),
            ]
            for k, (val, fmt, font) in enumerate(row_data):
                c = ws.cell(row=row, column=2+k, value=val)
                c.font = font; c.fill = fill
                if fmt:
                    c.number_format = fmt
                c.alignment = right if k > 0 else Alignment(vertical="center")

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()


# ── Main app ──────────────────────────────────────────────────────────────────
def main():
    inject_css()

    with st.sidebar:
        st.markdown(f"### 📊 Executive Dashboard")
        uploaded = st.file_uploader("Upload P&L Workbook (.xlsx)", type=["xlsx"])
        if uploaded:
            st.success(f"✅ {uploaded.name}")

    if not uploaded:
        st.markdown(f"""
        <div style="text-align:center; padding:80px 20px;">
            <h1 style="color:{COLORS['accent']}; font-size:36px;">📊 Executive Dashboard</h1>
            <p style="color:{COLORS['text_muted']}; font-size:16px; max-width:600px; margin:20px auto;">
                Upload a P&L workbook to generate your executive dashboard.
            </p>
            <p style="color:{COLORS['text_dim']}; font-size:13px;">
                ← Upload a .xlsx file in the sidebar to get started
            </p>
        </div>
        """, unsafe_allow_html=True)
        return

    file_bytes = uploaded.read()
    wb_tmp = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    sheet_names = wb_tmp.sheetnames
    wb_tmp.close()

    with st.sidebar:
        default_idx = 0
        for i, name in enumerate(sheet_names):
            nl = name.lower()
            if "consolidated" in nl and "p&l" in nl:
                default_idx = i; break
            elif "consolidated" in nl:
                default_idx = i
        selected_sheet = st.selectbox("Consolidated P&L Sheet", sheet_names, index=default_idx)

    df = pd.read_excel(io.BytesIO(file_bytes), sheet_name=selected_sheet, header=None)
    data = discover_and_extract(df)
    fy_list = data["fy_labels"]

    if not fy_list:
        st.error("Could not detect fiscal year columns. Check that your P&L has columns labeled 'FY20XX Total' or similar.")
        return

    latest_fy = fy_list[-1]
    prior_fy  = fy_list[-2] if len(fy_list) >= 2 else fy_list[0]
    latest    = data["annual"].get(latest_fy, {})
    prior     = data["annual"].get(prior_fy,  {})

    # Area sheets
    area_sheet_names = discover_area_sheets(sheet_names, selected_sheet)
    area_datasets = []
    for an in area_sheet_names:
        try:
            adf  = pd.read_excel(io.BytesIO(file_bytes), sheet_name=an, header=None)
            adata = discover_and_extract(adf)
            if adata["annual"]:
                display_name = an
                for s in ["_P&L", "_PL", " P&L", " PL"]:
                    display_name = display_name.replace(s, "")
                adata["display_name"] = display_name.strip()
                area_datasets.append(adata)
        except Exception:
            pass
    area_datasets.sort(
        key=lambda x: x["annual"].get(latest_fy, {}).get("total_revenue", 0), reverse=True
    )

    with st.sidebar:
        st.markdown("---")
        st.markdown("**Discovered Structure**")
        st.markdown(
            f'<p style="color:{COLORS["text_dim"]};font-size:11px;">'
            f'Fiscal Years: {", ".join(fy_list)}<br>'
            f'Service Lines: {len(data["service_lines"])}<br>'
            f'Monthly Data: {len(data["monthly_revenue"])} FYs<br>'
            f'Areas Found: {len(area_datasets)}</p>',
            unsafe_allow_html=True
        )

    # Build tabs
    tab_labels = ["📊 Dashboard", "📈 Trends"]
    if area_datasets:
        tab_labels.append("🗺️ Areas")
    tab_labels += ["🎯 Scenarios", "⬇️ Export"]
    tabs = st.tabs(tab_labels)
    tab_idx = 0

    # ── Dashboard tab ─────────────────────────────────────────────────────────
    with tabs[tab_idx]:
        tab_idx += 1
        rev      = latest.get("total_revenue", 0)
        prior_rev = prior.get("total_revenue", 1) or 1
        growth   = rev / prior_rev - 1

        kpi_items = [
            ("Total Revenue",       fmt_dollars(rev),          f"PY: {fmt_dollars(prior.get('total_revenue',0))}", COLORS["accent"]),
            ("Rev Growth YoY",      fmt_growth(growth),        f"PY GM%: {fmt_pct(prior.get('gm_pct',0))}",       COLORS["positive"]),
            ("Gross Margin %",      fmt_pct(latest.get("gm_pct",0)), f"PY: {fmt_pct(prior.get('gm_pct',0))}",    COLORS["warning"]),
            ("Contribution Margin", fmt_dollars(latest.get("cm",0)), f"PY: {fmt_dollars(prior.get('cm',0))}",     COLORS["accent_alt"]),
            ("CM %",                fmt_pct(latest.get("cm_pct",0)), f"PY: {fmt_pct(prior.get('cm_pct',0))}",    COLORS["negative"]),
        ]
        cols = st.columns(5)
        for col, (label, value, sub, bc) in zip(cols, kpi_items):
            with col:
                st.markdown(kpi_card(label, value, sub, bc), unsafe_allow_html=True)

        st.markdown("")

        # Service line table
        if data["service_lines"]:
            st.markdown(f'<div class="section-title">SERVICE LINE PERFORMANCE — {latest_fy} vs {prior_fy}</div>', unsafe_allow_html=True)
            sl_df = pd.DataFrame(data["service_lines"])
            if prior_fy in sl_df.columns and latest_fy in sl_df.columns:
                disp = sl_df[["name", prior_fy, latest_fy]].copy()
                disp.columns = ["Service Line", prior_fy, latest_fy]
                disp[prior_fy]  /= 1000
                disp[latest_fy] /= 1000
                disp["Variance"] = disp[latest_fy] - disp[prior_fy]
                disp["Var %"]    = ((disp["Variance"] / disp[prior_fy]) * 100).map(lambda x: f"{x:+.1f}%")
                st.dataframe(disp, use_container_width=True, hide_index=True)

        # Annual summary
        st.markdown(f'<div class="section-title">ANNUAL FINANCIAL SUMMARY</div>', unsafe_allow_html=True)
        summary_rows = []
        for label, key in [
            ("Total Revenue","total_revenue"),("Cost of Services","total_cos"),
            ("Gross Margin","total_gm"),("GM %","gm_pct"),("Direct Costs","total_direct"),
            ("Contribution Margin","cm"),("CM %","cm_pct"),("Operating Income","oi"),
        ]:
            row = {"Metric": label}
            for fy in fy_list:
                v = data["annual"].get(fy, {}).get(key, 0)
                row[fy] = f"{v*100:.1f}%" if "pct" in key else fmt_dollars(v)
            summary_rows.append(row)
        st.dataframe(pd.DataFrame(summary_rows), use_container_width=True, hide_index=True)

        # Revenue mix charts
        st.markdown(f'<div class="section-title">REVENUE MIX</div>', unsafe_allow_html=True)
        c1, c2 = st.columns(2)
        for col, fy in [(c1, prior_fy), (c2, latest_fy)]:
            sls = [sl for sl in data["service_lines"] if sl.get(fy, 0) > 0]
            if sls:
                fig = go.Figure(data=[go.Pie(
                    labels=[s["name"] for s in sls],
                    values=[s[fy]/1000 for s in sls],
                    hole=0.45, textinfo="percent",
                    textfont=dict(color="white", size=11),
                    marker=dict(colors=CHART_COLORS[:len(sls)])
                )])
                fig.update_layout(**plotly_base(f"Revenue Mix — {fy}", height=300))
                with col:
                    st.plotly_chart(fig, use_container_width=True)

        # P&L bridge + cost efficiency
        st.markdown(f'<div class="section-title">P&L BRIDGE &amp; COST EFFICIENCY</div>', unsafe_allow_html=True)
        c1, c2 = st.columns(2)
        fys3 = fy_list[-3:]
        bridge_fig = go.Figure()
        for i, fy in enumerate(fys3):
            a = data["annual"].get(fy, {})
            bridge_fig.add_trace(go.Bar(
                x=["Revenue", "Gross Margin", "CM"],
                y=[a.get("total_revenue",0)/1000, a.get("total_gm",0)/1000, a.get("cm",0)/1000],
                name=fy, marker_color=CHART_COLORS[2+i]
            ))
        bridge_fig.update_layout(**plotly_base("P&L Bridge"))
        bridge_fig.update_layout(barmode="group")
        with c1:
            st.plotly_chart(bridge_fig, use_container_width=True)

        eff_fig = go.Figure()
        for i, fy in enumerate(fys3):
            a = data["annual"].get(fy, {})
            r = a.get("total_revenue", 1) or 1
            eff_fig.add_trace(go.Bar(
                x=["COS %", "Direct %", "Total %"],
                y=[a.get("total_cos",0)/r*100, a.get("total_direct",0)/r*100,
                   (a.get("total_cos",0)+a.get("total_direct",0))/r*100],
                name=fy, marker_color=CHART_COLORS[2+i]
            ))
        eff_fig.update_layout(**plotly_base("Cost as % of Revenue"))
        eff_fig.update_layout(barmode="group")
        eff_fig.update_yaxes(ticksuffix="%")
        with c2:
            st.plotly_chart(eff_fig, use_container_width=True)

    # ── Trends tab ────────────────────────────────────────────────────────────
    with tabs[tab_idx]:
        tab_idx += 1
        months = data["month_labels"]

        st.markdown(f'<div class="section-title">MONTHLY REVENUE TREND</div>', unsafe_allow_html=True)
        fig = go.Figure()
        for i, (fy, vals) in enumerate(sorted(data["monthly_revenue"].items())):
            fig.add_trace(go.Scatter(
                x=months, y=[v/1000 for v in vals], name=fy,
                line=dict(color=CHART_COLORS[i % len(CHART_COLORS)], width=2),
                mode="lines+markers", marker=dict(size=4)
            ))
        fig.update_layout(**plotly_base("Monthly Revenue Trend", height=320))
        st.plotly_chart(fig, use_container_width=True)

        c1, c2 = st.columns(2)
        for col, metric_data, title, ylabel in [
            (c1, data["monthly_gm_pct"],  "Gross Margin % Trend",        "GM %"),
            (c2, data["monthly_cm_pct"],  "Contribution Margin % Trend", "CM %"),
        ]:
            fig = go.Figure()
            for i, (fy, vals) in enumerate(sorted(metric_data.items())):
                pvals = [v*100 if abs(v)<1 else v for v in vals]
                fig.add_trace(go.Scatter(
                    x=months, y=pvals, name=fy,
                    line=dict(color=CHART_COLORS[i % len(CHART_COLORS)], width=2),
                    mode="lines+markers", marker=dict(size=4)
                ))
            fig.update_layout(**plotly_base(title))
            fig.update_yaxes(title_text=ylabel, ticksuffix="%")
            with col:
                st.plotly_chart(fig, use_container_width=True)

        st.markdown(f'<div class="section-title">BUDGET vs PRIOR YEAR — MONTHLY</div>', unsafe_allow_html=True)
        fig = go.Figure()
        for i, fy in enumerate(fy_list[-3:]):
            vals = data["monthly_revenue"].get(fy, [])
            fig.add_trace(go.Bar(
                x=months, y=[v/1000 for v in vals], name=fy,
                marker_color=CHART_COLORS[2 + i]
            ))
        fig.update_layout(**plotly_base("Monthly Revenue — Budget vs Prior Year", height=320))
        fig.update_layout(barmode="group")
        st.plotly_chart(fig, use_container_width=True)

    # ── Areas tab ─────────────────────────────────────────────────────────────
    if area_datasets:
        with tabs[tab_idx]:
            tab_idx += 1
            st.markdown(f'<div class="section-title">AREA PERFORMANCE OVERVIEW — {len(area_datasets)} Areas</div>', unsafe_allow_html=True)

            cols_per_row = min(len(area_datasets), 4)
            for row_start in range(0, len(area_datasets), cols_per_row):
                cols = st.columns(cols_per_row)
                for i, col in enumerate(cols):
                    idx = row_start + i
                    if idx >= len(area_datasets):
                        break
                    ad = area_datasets[idx]
                    al = ad["annual"].get(latest_fy, {})
                    ap = ad["annual"].get(prior_fy,  {})
                    rl = al.get("total_revenue", 0)
                    rp = ap.get("total_revenue", 1) or 1
                    g  = (rl / rp - 1) * 100
                    with col:
                        st.markdown(
                            area_card(ad["display_name"], rl, al.get("cm_pct",0), g, AREA_COLORS[idx % len(AREA_COLORS)]),
                            unsafe_allow_html=True
                        )

            st.markdown("")
            c1, c2 = st.columns(2)

            # Revenue comparison
            area_names = [ad["display_name"] for ad in area_datasets]
            rev_fig = go.Figure()
            for i, fy in enumerate(fy_list[-3:]):
                rev_fig.add_trace(go.Bar(
                    x=area_names,
                    y=[ad["annual"].get(fy, {}).get("total_revenue", 0)/1000 for ad in area_datasets],
                    name=fy, marker_color=CHART_COLORS[2+i]
                ))
            rev_fig.update_layout(**plotly_base("Revenue by Area", height=380))
            rev_fig.update_layout(barmode="group")
            with c1:
                st.plotly_chart(rev_fig, use_container_width=True)

            # Revenue share doughnut
            share_vals = [ad["annual"].get(latest_fy, {}).get("total_revenue", 0)/1000 for ad in area_datasets]
            share_fig = go.Figure(data=[go.Pie(
                labels=area_names, values=share_vals, hole=0.45,
                textinfo="percent+label", textfont=dict(color="white", size=10),
                marker=dict(colors=AREA_COLORS[:len(area_names)])
            )])
            share_fig.update_layout(**plotly_base(f"Revenue Share — {latest_fy}", height=380))
            with c2:
                st.plotly_chart(share_fig, use_container_width=True)

            c1, c2 = st.columns(2)

            # CM% comparison
            cm_fig = go.Figure()
            for i, fy in enumerate(fy_list[-3:]):
                cm_fig.add_trace(go.Bar(
                    x=area_names,
                    y=[ad["annual"].get(fy, {}).get("cm_pct", 0)*100 for ad in area_datasets],
                    name=fy, marker_color=CHART_COLORS[2+i]
                ))
            cm_fig.update_layout(**plotly_base("CM % by Area", height=380))
            cm_fig.update_layout(barmode="group")
            cm_fig.update_yaxes(ticksuffix="%")
            with c1:
                st.plotly_chart(cm_fig, use_container_width=True)

            # Revenue growth
            g_vals = []
            g_cols = []
            for ad in area_datasets:
                rl = ad["annual"].get(latest_fy, {}).get("total_revenue", 0)
                rp = ad["annual"].get(prior_fy,  {}).get("total_revenue", 1) or 1
                g = (rl / rp - 1) * 100
                g_vals.append(g)
                g_cols.append(COLORS["positive"] if g >= 0 else COLORS["negative"])
            grow_fig = go.Figure(data=[go.Bar(
                x=area_names, y=g_vals, marker_color=g_cols,
                text=[f"{v:+.1f}%" for v in g_vals], textposition="auto",
                textfont=dict(color="white", size=11)
            )])
            grow_fig.update_layout(**plotly_base(f"Revenue Growth — {prior_fy} → {latest_fy}", height=380))
            grow_fig.update_yaxes(ticksuffix="%")
            with c2:
                st.plotly_chart(grow_fig, use_container_width=True)

            # Area table
            st.markdown(f'<div class="section-title">AREA PERFORMANCE TABLE</div>', unsafe_allow_html=True)
            area_rows = []
            for ad in area_datasets:
                ap = ad["annual"].get(prior_fy,  {})
                al = ad["annual"].get(latest_fy, {})
                rp = ap.get("total_revenue", 0) / 1000
                rl = al.get("total_revenue", 0) / 1000
                g  = (rl / rp - 1) * 100 if rp else 0
                first_fy = sorted(ad["annual"].keys())[0]
                rf = ad["annual"].get(first_fy, {}).get("total_revenue", 0) / 1000
                n  = max(1, int(latest_fy[2:]) - int(first_fy[2:]))
                cagr = ((rl / rf) ** (1/n) - 1) * 100 if rf > 0 else 0
                area_rows.append({
                    "Area":              ad["display_name"],
                    f"Rev {prior_fy}":   fmt_dollars(rp*1000),
                    f"Rev {latest_fy}":  fmt_dollars(rl*1000),
                    "Growth %":          f"{g:+.1f}%",
                    f"GM% {latest_fy}":  fmt_pct(al.get("gm_pct",0)),
                    f"CM% {latest_fy}":  fmt_pct(al.get("cm_pct",0)),
                    "CAGR":              f"{cagr:.1f}%",
                })
            st.dataframe(pd.DataFrame(area_rows), use_container_width=True, hide_index=True)

    # ── Scenarios tab ─────────────────────────────────────────────────────────
    with tabs[tab_idx]:
        tab_idx += 1
        st.markdown(f'<div class="section-title">SCENARIO ANALYSIS — {latest_fy}</div>', unsafe_allow_html=True)

        c1, c2, c3 = st.columns(3)
        with c1:
            st.markdown("**▲ Bull Case**")
            bull_rev = st.slider("Revenue Growth (Bull)", -20.0, 30.0,  5.0, 0.5, key="bull_rev") / 100
            bull_cos = st.slider("COS % Adj (Bull)",     -10.0, 10.0, -2.0, 0.5, key="bull_cos") / 100
        with c2:
            st.markdown("**● Base Case**")
            st.info(f"Revenue: {fmt_dollars(latest.get('total_revenue',0))}\n\nCM%: {fmt_pct(latest.get('cm_pct',0))}")
        with c3:
            st.markdown("**▼ Bear Case**")
            bear_rev = st.slider("Revenue Growth (Bear)", -30.0, 10.0, -10.0, 0.5, key="bear_rev") / 100
            bear_cos = st.slider("COS % Adj (Bear)",       -5.0, 15.0,   3.0, 0.5, key="bear_cos") / 100

        bull = compute_scenario(latest, bull_rev, bull_cos)
        base = compute_scenario(latest, 0, 0)
        bear = compute_scenario(latest, bear_rev, bear_cos)

        scen_df = pd.DataFrame([
            {"Scenario":"▲ Bull", "Revenue":fmt_dollars(bull["revenue"]*1000), "GM":fmt_dollars(bull["gm"]*1000), "CM":fmt_dollars(bull["cm"]*1000), "CM %":fmt_pct(bull["cm_pct"])},
            {"Scenario":"● Base", "Revenue":fmt_dollars(base["revenue"]*1000), "GM":fmt_dollars(base["gm"]*1000), "CM":fmt_dollars(base["cm"]*1000), "CM %":fmt_pct(base["cm_pct"])},
            {"Scenario":"▼ Bear", "Revenue":fmt_dollars(bear["revenue"]*1000), "GM":fmt_dollars(bear["gm"]*1000), "CM":fmt_dollars(bear["cm"]*1000), "CM %":fmt_pct(bear["cm_pct"])},
        ])
        st.dataframe(scen_df, use_container_width=True, hide_index=True)

        fig = go.Figure()
        for s, color, label in [(bull, COLORS["positive"],"Bull"), (base, COLORS["accent_alt"],"Base"), (bear, COLORS["negative"],"Bear")]:
            fig.add_trace(go.Bar(
                x=["Revenue","Gross Margin","Contribution Margin"],
                y=[s["revenue"], s["gm"], s["cm"]],
                name=label, marker_color=color
            ))
        fig.update_layout(**plotly_base("Scenario Comparison"))
        fig.update_layout(barmode="group")
        st.plotly_chart(fig, use_container_width=True)

        # Sensitivity matrix heatmap
        st.markdown(f'<div class="section-title">CM % SENSITIVITY — Revenue Growth vs COS Adjustment</div>', unsafe_allow_html=True)
        rev_steps = [-0.10, -0.05, 0, 0.05, 0.10]
        cos_steps = [-0.03, -0.01,  0, 0.02,  0.05]
        matrix = [[compute_scenario(latest, r, c)["cm_pct"] * 100 for c in cos_steps] for r in rev_steps]
        heat_fig = go.Figure(data=go.Heatmap(
            z=matrix,
            x=[f"{c*100:+.0f}% COS" for c in cos_steps],
            y=[f"{r*100:+.0f}% Rev" for r in rev_steps],
            colorscale=[[0, COLORS["negative"]], [0.4, COLORS["warning"]], [0.6, COLORS["accent_alt"]], [1, COLORS["positive"]]],
            text=[[f"{v:.1f}%" for v in row] for row in matrix],
            texttemplate="%{text}", textfont=dict(size=12, color="white"),
        ))
        heat_fig.update_layout(**plotly_base("", height=360))
        heat_fig.update_layout(
            xaxis=dict(title="COS % Adjustment", side="top"),
            yaxis=dict(title="Revenue Growth Adjustment", autorange="reversed")
        )
        st.plotly_chart(heat_fig, use_container_width=True)

    # ── Export tab ────────────────────────────────────────────────────────────
    with tabs[tab_idx]:
        st.markdown(f'<div class="section-title">DOWNLOAD EXCEL DASHBOARD</div>', unsafe_allow_html=True)
        st.markdown(
            f'<p style="color:{COLORS["text_muted"]}">Download your original workbook with a new '
            f'<strong>Dashboard</strong> sheet prepended, including KPI scorecards, annual financial '
            f'summary, and area performance table.</p>',
            unsafe_allow_html=True
        )
        if st.button("🔧 Generate Excel Dashboard", type="primary"):
            with st.spinner("Building Excel dashboard..."):
                excel_bytes = generate_excel(data, area_datasets, file_bytes)
            st.download_button(
                label="⬇️ Download Dashboard Workbook",
                data=excel_bytes,
                file_name=f"Executive_Dashboard_{uploaded.name}",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            st.success("✅ Dashboard generated! Click above to download.")
        st.markdown(
            f'<p style="color:{COLORS["text_dim"]};font-size:11px;margin-top:20px;">'
            f'Export includes: KPI scorecards, annual financial summary, and area performance table.<br>'
            f'For fully formula-driven Excel dashboards with 16 charts, use the Claude Excel add-in.</p>',
            unsafe_allow_html=True
        )


if __name__ == "__main__":
    main()
